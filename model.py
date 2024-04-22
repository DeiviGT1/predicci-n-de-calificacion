from sklearn.neural_network import MLPClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.preprocessing import OneHotEncoder, MinMaxScaler
import pandas as pd
import time

from data_processing import DataProcessor
from data_preparation import carga_datos, procesamiento_tabla
from openpyxl import load_workbook

#################################################################### Descripción ########################################################################
#########################################################################################################################################################

# Este script se centra en el entrenamiento de modelos de machine learning utilizando diferentes clasificadores. Se encarga de 
# cargar los datos preparados, entrenar los modelos y generar predicciones para un conjunto específico de datos, con el fin de 
# facilitar la toma de decisiones basadas en el análisis de los resultados obtenidos.


########################################################################################################################################################
########################################################################################################################################################

ruta1 = "XXX"
ruta2 = "XXX"

# Clase para entrenar modelos
class ModelTrainer:
    def __init__(self, genero, mes_exh):
         # Definir clasificadores y diccionario de predicciones
        self.classifiers = [
            MLPClassifier(warm_start=True, random_state=5, validation_fraction=0.3, solver='adam', n_iter_no_change=20, max_iter=1000, learning_rate_init=0.1, learning_rate='adaptive', hidden_layer_sizes=(100,), early_stopping=True, alpha=0.01, activation='relu'),
            DecisionTreeClassifier(random_state=5, criterion='entropy', max_depth=10, max_features='sqrt', min_samples_leaf=4, min_samples_split=2, splitter='best'),
            KNeighborsClassifier(algorithm='auto', n_neighbors=40, p=2, weights='uniform')
        ]
        self.dict_pred = {}
        self.columns_to_remove = ['genero', 'categoria','categorizacion']
        if genero == 'hombre':
            self.columns_to_remove.extend(['base_tela', 'sub_categoria'])
        self.mes_exh = mes_exh
        
    
    
    # Método para entrenar modelos
    def train_models(self, X, y, items_despacho=None, prueba=False):
        for index, classi in enumerate(self.classifiers):
            classi_name = classi.__class__.__name__
            pipeline = DataProcessor.create_pipeline(cat_encoder=OneHotEncoder(), model=classi, num_encoder=MinMaxScaler())
            
            pipeline.fit(X, y)
            
            try:
                X_items = self.items_validator(items=items_despacho, prueba=prueba)
                self.dict_pred[classi_name] = pipeline.predict(X_items)
            except ValueError as e:
                # Extraer el nombre de la categoría desconocida y la columna correspondiente desde la excepción
                error_message = str(e)
                category = error_message.split("categories ")[1].split(" in column")[0]
                column = error_message.split("in column ")[1].split(" ")[0]
                
                # Imprimir el mensaje de error personalizado
                error_msg = f"\n Hay un error, corregir el dato {category} en la columna {X_items.columns[int(column)]} \n\n\n\n "
                raise ValueError(error_msg) from e
        
        for keys, values in self.dict_pred.items():
            self.t_despacho[keys] = values

        self.t_despacho.insert(0, 'genero', genero)
        self.t_despacho.insert(1, 'categoria', categoria)

        self.t_despacho.to_csv(f'{ruta1}_{genero}_{categoria}_{time.strftime("%Y-%m-%d_%H-%M-%S")}.csv', index=True)

        return self.dict_pred

    # Método para filtrar items
    def items_validator(self, items, prueba):
        if prueba:
            ws = wb['pruebas']
            self.t_despacho = pd.DataFrame(ws.values)
            self.t_despacho.columns = self.t_despacho.iloc[0]
            self.t_despacho = self.t_despacho[1:].dropna(subset=['item']).iloc[:, :10]
            self.t_despacho.index = self.t_despacho['item']
            self.t_despacho.drop(['item', 'genero', 'categoria'], axis=1, inplace=True)
            self.t_despacho = procesamiento_tabla(df=self.t_despacho)
            if genero == 'hombre':
                self.t_despacho.drop(['base_tela', 'sub_categoria'], axis=1, inplace=True)
              
            self.t_despacho['precio'] = self.t_despacho['precio'] / 1000 #Se agrega esta linea por la codificación en el maestro de ítems que toma los precios en miles
        else:
            self.t_despacho = carga_datos(mes_exh=self.mes_exh, columns_to_remove = self.columns_to_remove)
            self.t_despacho = self.t_despacho[self.t_despacho['item'].isin(items)]
            self.t_despacho.index = self.t_despacho['item']
            self.t_despacho = self.t_despacho.drop(columns=['item'])

        return self.t_despacho


if __name__ == "__main__":

    ################################################ PARAMETROS ################################################
    prueba = True
    #############################################################################################################
    
    wb = load_workbook(f'{ruta2}', data_only=True)
    ws = wb['pruebas']
    genero = ws['B2'].value
    categoria = ws['C2'].value
    mes_exh = ws['M1'].value
    items_despacho = [cell.value for cell in ws['A'] if cell.value != None]
    
    X_train, X_test, y_train, y_test, X, y = DataProcessor.data_final(genero, categoria, 7)
    
    model_trainer = ModelTrainer(genero, mes_exh)
    model_trainer.train_models(X, y, items_despacho, prueba)